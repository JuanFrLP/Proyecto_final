from openpyxl import Workbook, load_workbook
import customtkinter as ctk
from tkinter import messagebox, ttk
import os
from datetime import datetime

# CONFIGURACIÓN
NOMBRE_ARCHIVO = "inventario_hilos.xlsx"
PERMITIR_VENTA_EMPLEADO = True
STOCK_MINIMO = 6  # alerta cuando quede <= a este valor

# Nombres de hojas en Excel
HOJA_INVENTARIO = "Inventario_Hilos"
HOJA_COMPRAS = "Historial_Compras"
HOJA_VENTAS = "Historial_Ventas"
HOJA_USUARIOS = "Usuarios"
HOJA_SESIONES = "Sesiones_Usuarios"


# UTILIDADES
class Utilidades:
    @staticmethod
    def leer_entero_str(valor, minimo=None):
        try:
            n = int(valor)
            if minimo is not None and n < minimo:
                return None
            return n
        except Exception:
            return None

    @staticmethod
    def leer_float_str(valor, minimo=None):
        try:
            f = float(valor)
            if minimo is not None and f < minimo:
                return None
            return f
        except Exception:
            return None


# MANEJO DE EXCEL
class GestorExcel:
    def __init__(self, archivo):
        self.nombre_archivo = archivo

    def asegurar_estructura(self):
        if os.path.exists(self.nombre_archivo):
            wb = load_workbook(self.nombre_archivo)
        else:
            wb = Workbook()

        # Si el libro viene vacío, usa la hoja activa como Inventario
        if HOJA_INVENTARIO not in wb.sheetnames:
            hoja_inv = wb.active
            hoja_inv.title = HOJA_INVENTARIO
            hoja_inv.append(["ID", "Marca", "Código de Color", "Descripción", "Cantidad", "Precio Unitario", "Proveedor"])

        # Crear/asegurar el resto de hojas
        definiciones = {
            HOJA_COMPRAS: ["Código", "Marca", "Descripción", "Cantidad", "Costo Unitario", "Total"],
            HOJA_VENTAS: ["Código", "Marca", "Descripción", "Cantidad", "Total"],
            HOJA_USUARIOS: ["Usuario", "Contraseña", "Rol"],
            HOJA_SESIONES: ["ID Sesión", "Usuario", "Rol", "Fecha y Hora de Inicio", "Fecha y Hora de Cierre"],
        }

        for hoja, encabezados in definiciones.items():
            if hoja not in wb.sheetnames:
                nueva = wb.create_sheet(hoja)
                nueva.append(encabezados)
                if hoja == HOJA_USUARIOS:
                    # usuarios por defecto
                    nueva.append(["admin", "admin123", "admin"])
                    nueva.append(["empleado", "azul321", "user"])
            else:
                obj = wb[hoja]
                if obj.max_row == 0:
                    obj.append(encabezados)

        wb.save(self.nombre_archivo)

    def cargar_hoja(self, nombre_hoja):
        if not os.path.exists(self.nombre_archivo):
            self.asegurar_estructura()
        wb = load_workbook(self.nombre_archivo)
        if nombre_hoja not in wb.sheetnames:
            return []
        hoja = wb[nombre_hoja]
        return [fila for fila in hoja.iter_rows(min_row=2, values_only=True) if fila and fila[0] is not None]

    def guardar_hoja(self, nombre_hoja, encabezados, datos):
        if os.path.exists(self.nombre_archivo):
            wb = load_workbook(self.nombre_archivo)
        else:
            wb = Workbook()
        if nombre_hoja not in wb.sheetnames:
            hoja = wb.create_sheet(nombre_hoja)
            hoja.append(encabezados)
        else:
            hoja = wb[nombre_hoja]
            if hoja.max_row > 1:
                hoja.delete_rows(2, hoja.max_row - 1)
        for fila in datos:
            hoja.append(fila)
        wb.save(self.nombre_archivo)

    def reparar_estructura(self):
        if not os.path.exists(self.nombre_archivo):
            return

        wb = load_workbook(self.nombre_archivo)

        hojas_principales = [HOJA_INVENTARIO, HOJA_COMPRAS, HOJA_VENTAS]
        encabezados_validos = {
            HOJA_INVENTARIO: ["ID", "Marca", "Código de Color", "Descripción", "Cantidad", "Precio Unitario", "Proveedor"],
            HOJA_COMPRAS: ["Código", "Marca", "Descripción", "Cantidad", "Costo Unitario", "Total"],
            HOJA_VENTAS: ["Código", "Marca", "Descripción", "Cantidad", "Total"],
        }

        for hoja_nombre in hojas_principales:
            if hoja_nombre not in wb.sheetnames:
                continue

            hoja = wb[hoja_nombre]
            # Detectar encabezados incorrectos en filas posteriores
            filas_erroneas = []
            for i, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
                if fila and any(isinstance(v, str) and v.strip() in encabezados_validos[hoja_nombre] for v in fila):
                    if i != 1:  # Si no es la primera fila (encabezado correcto)
                        filas_erroneas.append(i)

            # Borrar filas con encabezados repetidos
            if filas_erroneas:
                hoja.delete_rows(2, hoja.max_row - 1)

        wb.save(self.nombre_archivo)


# SESIONES
class SesionUsuario:
    def __init__(self, gestor_excel: GestorExcel):
        self.excel = gestor_excel

    def abrir_sesion(self, usuario, rol):
        wb = load_workbook(self.excel.nombre_archivo)
        if HOJA_SESIONES not in wb.sheetnames:
            hoja = wb.create_sheet(HOJA_SESIONES)
            hoja.append(["ID Sesión", "Usuario", "Rol", "Fecha y Hora de Inicio", "Fecha y Hora de Cierre"])
        hoja = wb[HOJA_SESIONES]
        id_sesion = hoja.max_row  # correlativo
        inicio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        hoja.append([id_sesion, usuario, rol, inicio, ""])
        wb.save(self.excel.nombre_archivo)
        return id_sesion

    def cerrar_sesion(self, id_sesion):
        wb = load_workbook(self.excel.nombre_archivo)
        if HOJA_SESIONES not in wb.sheetnames:
            return
        hoja = wb[HOJA_SESIONES]
        cierre = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for fila in hoja.iter_rows(min_row=2):
            if fila[0].value == id_sesion:
                hoja.cell(row=fila[0].row, column=5, value=cierre)
                break
        wb.save(self.excel.nombre_archivo)

# INVENTARIO / COMPRAS / VENTAS / REPORTES
class InventarioHilos:
    def __init__(self, gestor_excel: GestorExcel, stock_minimo=STOCK_MINIMO):
        self.excel = gestor_excel
        self.stock_minimo = stock_minimo
        self.inventario = []
        self.historial_ventas = []
        self.historial_compras = []
        self.contador_id = 1
        # --- Índice Hash: (marca.lower(), codigo_str) -> dict del hilo
        self.hash_index = {}
        self.cargar_todo()

    # ---- Utilidad interna: reconstruir el índice hash
    def _rebuild_hash_index(self):
        self.hash_index.clear()
        for h in self.inventario:
            clave = (h["marca"].lower(), str(h["codigo_color"]))
            self.hash_index[clave] = h

    def cargar_todo(self):
        self.inventario.clear()
        self.historial_ventas.clear()
        self.historial_compras.clear()

        # --- Inventario ---
        for fila in self.excel.cargar_hoja(HOJA_INVENTARIO):
            try:
                id_val = int(fila[0])
            except (ValueError, TypeError):
                continue
            try:
                cantidad_val = int(fila[4])
            except (ValueError, TypeError):
                cantidad_val = 0
            try:
                precio_val = float(fila[5])
            except (ValueError, TypeError):
                precio_val = 0.0

            self.inventario.append(
                {
                    "id": id_val,
                    "marca": fila[1],
                    "codigo_color": str(fila[2]),
                    "descripcion": fila[3],
                    "cantidad": cantidad_val,
                    "precio_unitario": precio_val,
                    "proveedor": fila[6],
                }
            )

        # --- Historial de Compras ---
        for f in self.excel.cargar_hoja(HOJA_COMPRAS):
            try:
                cant = int(f[3])
                costo = float(f[4])
                total = float(f[5])
            except (ValueError, TypeError):
                continue
            self.historial_compras.append(
                {
                    "codigo_color": str(f[0]),
                    "marca": f[1],
                    "descripcion": f[2],
                    "cantidad": cant,
                    "costo_unitario": costo,
                    "total": total,
                }
            )

        # --- Historial de Ventas ---
        for f in self.excel.cargar_hoja(HOJA_VENTAS):
            try:
                cant = int(f[3])
                total = float(f[4])
            except (ValueError, TypeError):
                continue
            self.historial_ventas.append(
                {
                    "codigo_color": str(f[0]),
                    "marca": f[1],
                    "descripcion": f[2],
                    "cantidad": cant,
                    "total": total,
                }
            )

        # --- Asignar ID ---
        self.contador_id = max((h["id"] for h in self.inventario), default=0) + 1

        # --- Construir índice hash
        self._rebuild_hash_index()

        # --- Ordenar automáticamente y guardar (si hay inventario)
        if self.inventario:
            self.inventario = self.ordenar_por_codigo_color(self.inventario)
            self.inventario = self.ordenar_por_marca_con_menos_stock(self.inventario)
            self.inventario = self.ordenar_por_tipo(self.inventario)
            self.guardar_todo()

    def guardar_todo(self):
        datos_inv = [
            [h["id"], h["marca"], h["codigo_color"], h["descripcion"], h["cantidad"], h["precio_unitario"], h["proveedor"]]
            for h in self.inventario
        ]
        datos_comp = [
            [c["codigo_color"], c["marca"], c["descripcion"], c["cantidad"], c["costo_unitario"], c["total"]]
            for c in self.historial_compras
        ]
        datos_vent = [
            [v["codigo_color"], v["marca"], v["descripcion"], v["cantidad"], v["total"]]
            for v in self.historial_ventas
        ]

        self.excel.guardar_hoja(
            HOJA_INVENTARIO,
            ["ID", "Marca", "Código de Color", "Descripción", "Cantidad", "Precio Unitario", "Proveedor"],
            datos_inv,
        )
        self.excel.guardar_hoja(
            HOJA_COMPRAS, ["Código", "Marca", "Descripción", "Cantidad", "Costo Unitario", "Total"], datos_comp
        )
        self.excel.guardar_hoja(HOJA_VENTAS, ["Código", "Marca", "Descripción", "Cantidad", "Total"], datos_vent)

    # ---- Alertas ----
    def avisar_stock_bajo_gui(self, hilo):
        if hilo["cantidad"] <= self.stock_minimo:
            messagebox.showwarning(
                "Stock bajo",
                f"Código: {hilo['codigo_color']}\n"
                f"Tipo: {hilo['descripcion']}\n"
                f"Unidades: {hilo['cantidad']}",
            )

    # ---- Apoyo de negocio (sin inputs/prints) ----
    def existe_marca_codigo(self, marca, codigo):  # Búsqueda secuencial
        return any(
            h["marca"].lower() == marca.lower() and str(h["codigo_color"]) == str(codigo) for h in self.inventario
        )

    def obtener_por_marca_codigo(self, marca, codigo):  # Búsqueda secuencial
        for h in self.inventario:
            if h["marca"].lower() == marca.lower() and str(h["codigo_color"]) == str(codigo):
                return h
        return None

    # ---- Búsqueda por HASH (O(1) promedio) ----
    def buscar_hash(self, marca, codigo):
        clave = (marca.lower(), str(codigo))
        return self.hash_index.get(clave, None)

    # ---- CRUD de Hilos (usados por la GUI) ----
    def registrar_hilo_gui(self, marca, tipo, codigo_color, cantidad, precio_unitario, proveedor):
        if self.existe_marca_codigo(marca, codigo_color):
            return False, "Ese código de color ya existe para esta marca."

        hilo = {
            "id": self.contador_id,
            "marca": marca,
            "codigo_color": str(codigo_color),
            "descripcion": tipo,  # mapeo UI "Tipo" -> Excel "Descripción"
            "cantidad": cantidad,
            "precio_unitario": precio_unitario,
            "proveedor": proveedor,
        }
        self.inventario.append(hilo)
        self.contador_id += 1

        # Actualizar hash (clave nueva)
        self.hash_index[(hilo["marca"].lower(), str(hilo["codigo_color"]))] = hilo

        # Reordenamientos y guardado
        self.inventario = self.ordenar_por_codigo_color(self.inventario)
        self.inventario = self.ordenar_por_marca_con_menos_stock(self.inventario)
        self.inventario = self.ordenar_por_tipo(self.inventario)
        self.guardar_todo()
        return True, f"Hilo guardado. ID: {hilo['id']}"

    def modificar_hilo_gui(
        self,
        marca,
        codigo_color,
        nueva_marca=None,
        nuevo_tipo=None,
        nueva_cantidad=None,
        nuevo_precio=None,
        nuevo_proveedor=None,
    ):
        h = self.obtener_por_marca_codigo(marca, codigo_color)
        if not h:
            return False, "No se encontró el hilo con esa Marca y Código."

        # Si cambia la marca y/o el código, validar que no choque con otro existente
        destino_marca = nueva_marca.strip() if (nueva_marca is not None and nueva_marca.strip() != "") else h["marca"]
        destino_codigo = str(codigo_color)  # no estamos cambiando código en este formulario

        if (destino_marca.lower() != h["marca"].lower()) and self.existe_marca_codigo(destino_marca, destino_codigo):
            return False, "Ya existe ese código en la nueva marca."

        # Antes: clave vieja
        clave_vieja = (h["marca"].lower(), str(h["codigo_color"]))

        h["marca"] = destino_marca
        if (nuevo_tipo is not None) and nuevo_tipo.strip() != "":
            h["descripcion"] = nuevo_tipo.strip()
        if (nueva_cantidad is not None) and nueva_cantidad != "":
            n = Utilidades.leer_entero_str(nueva_cantidad, minimo=0)
            if n is None:
                return False, "Cantidad inválida."
            h["cantidad"] = n
            self.avisar_stock_bajo_gui(h)
        if (nuevo_precio is not None) and nuevo_precio != "":
            f = Utilidades.leer_float_str(nuevo_precio, minimo=0.0)
            if f is None:
                return False, "Precio inválido."
            h["precio_unitario"] = f
        if (nuevo_proveedor is not None) and nuevo_proveedor.strip() != "":
            h["proveedor"] = nuevo_proveedor.strip()

        # Después: actualizar hash (puede cambiar la marca)
        self._rebuild_hash_index()

        # Ordenar/guardar
        self.inventario = self.ordenar_por_codigo_color(self.inventario)
        self.inventario = self.ordenar_por_marca_con_menos_stock(self.inventario)
        self.inventario = self.ordenar_por_tipo(self.inventario)
        self.guardar_todo()
        return True, "Cambios guardados."

    def eliminar_hilo_gui(self, marca, codigo_color):
        h = self.obtener_por_marca_codigo(marca, codigo_color)
        if not h:
            return False, "No se encontró el hilo con esa Marca y Código."
        if h["cantidad"] != 0:
            return False, "No se puede eliminar: aún hay unidades."
        self.inventario.remove(h)

        # Actualizar hash
        clave = (h["marca"].lower(), str(h["codigo_color"]))
        if clave in self.hash_index:
            del self.hash_index[clave]

        self.guardar_todo()
        return True, "Hilo eliminado."

    def registrar_compra_gui(self, marca, codigo_color, cantidad, costo_unitario):
        h = self.obtener_por_marca_codigo(marca, codigo_color)
        if not h:
            return False, "No se encontró el hilo con esa Marca y Código."
        if cantidad < 1:
            return False, "Cantidad inválida."

        h["cantidad"] += cantidad
        self.historial_compras.append(
            {
                "codigo_color": str(codigo_color),
                "marca": h["marca"],
                "descripcion": h["descripcion"],
                "cantidad": cantidad,
                "costo_unitario": costo_unitario,
                "total": round(cantidad * costo_unitario, 2),
            }
        )
        self.avisar_stock_bajo_gui(h)

        # Ordenar/guardar
        self.inventario = self.ordenar_por_codigo_color(self.inventario)
        self.inventario = self.ordenar_por_marca_con_menos_stock(self.inventario)
        self.inventario = self.ordenar_por_tipo(self.inventario)
        self.guardar_todo()
        return True, "Compra registrada y stock actualizado."

    def registrar_venta_gui(self, marca, codigo_color, cantidad):
        h = self.obtener_por_marca_codigo(marca, codigo_color)
        if not h:
            return False, "No se encontró el hilo con esa Marca y Código."
        if cantidad < 1:
            return False, "Cantidad inválida."
        if cantidad > h["cantidad"]:
            return False, "No alcanza el stock para esa venta."

        h["cantidad"] -= cantidad
        total = round(cantidad * h["precio_unitario"], 2)
        self.historial_ventas.append(
            {
                "codigo_color": str(codigo_color),
                "marca": h["marca"],
                "descripcion": h["descripcion"],
                "cantidad": cantidad,
                "total": total,
            }
        )
        self.avisar_stock_bajo_gui(h)

        # Ordenar/guardar
        self.inventario = self.ordenar_por_codigo_color(self.inventario)
        self.inventario = self.ordenar_por_marca_con_menos_stock(self.inventario)
        self.inventario = self.ordenar_por_tipo(self.inventario)
        self.guardar_todo()
        return True, f"Venta registrada. Total: Q{total:.2f}"

    # ---- Reportes simples (devuelven listas) ----
    def reporte_inventario(self):
        return list(self.inventario)

    def reporte_ventas(self):
        return list(self.historial_ventas)

    def reporte_compras(self):
        return list(self.historial_compras)

    # ---- ORDENAMIENTOS ----
    def ordenar_por_codigo_color(self, inventario):  # QuickSort por código dentro de cada marca
        if not inventario:
            return inventario

        grupos = {}
        for hilo in inventario:
            marca = hilo["marca"]
            grupos.setdefault(marca, []).append(hilo)

        inventario_ordenado = []
        for marca, lista in sorted(grupos.items(), key=lambda x: x[0].lower()):
            lista_ordenada = self._quick_sort_codigo(lista)
            inventario_ordenado.extend(lista_ordenada)
        return inventario_ordenado

    def _quick_sort_codigo(self, lista):  # QuickSort
        if len(lista) <= 1:
            return lista
        pivote = lista[len(lista) // 2]
        try:
            p_color = int(pivote["codigo_color"])
        except ValueError:
            p_color = 0

        menores, iguales, mayores = [], [], []
        for x in lista:
            try:
                color = int(x["codigo_color"])
            except ValueError:
                color = 0
            if color < p_color:
                menores.append(x)
            elif color == p_color:
                iguales.append(x)
            else:
                mayores.append(x)

        return self._quick_sort_codigo(menores) + iguales + self._quick_sort_codigo(mayores)

    def ordenar_por_marca_con_menos_stock(self, inventario):  # Selection sort por stock total de marca (asc)
        resumen = {}
        for hilo in inventario:
            marca = hilo["marca"]
            cantidad = int(hilo["cantidad"])
            resumen[marca] = resumen.get(marca, 0) + cantidad

        lista_aux = [{**h, "total_marca": resumen[h["marca"]]} for h in inventario]

        n = len(lista_aux)
        for i in range(n):
            min_idx = i
            for j in range(i + 1, n):
                if lista_aux[j]["total_marca"] < lista_aux[min_idx]["total_marca"]:
                    min_idx = j
            lista_aux[i], lista_aux[min_idx] = lista_aux[min_idx], lista_aux[i]

        for h in lista_aux:
            del h["total_marca"]

        return lista_aux

    def ordenar_por_tipo(self, inventario):  # Shell Sort por descripción (tipo)
        lista_ordenada = inventario[:]
        n = len(lista_ordenada)
        gap = n // 2
        while gap > 0:
            for i in range(gap, n):
                temp = lista_ordenada[i]
                j = i
                while (
                    j >= gap
                    and lista_ordenada[j - gap]["descripcion"].lower() > temp["descripcion"].lower()
                ):
                    lista_ordenada[j] = lista_ordenada[j - gap]
                    j -= gap
                lista_ordenada[j] = temp
            gap //= 2
        return lista_ordenada
    
# SISTEMA (usuarios + sesión)
class SistemaDeInventario:
    def __init__(self, archivo_excel=NOMBRE_ARCHIVO, permitir_venta_empleado=PERMITIR_VENTA_EMPLEADO):
        self.excel = GestorExcel(archivo_excel)
        self.excel.asegurar_estructura()
        self.excel.reparar_estructura()
        self.inventario = InventarioHilos(self.excel, stock_minimo=STOCK_MINIMO)
        self.sesion = SesionUsuario(self.excel)
        self.usuarios = self._cargar_usuarios()
        self.permitir_venta_empleado = permitir_venta_empleado
        self.id_sesion_activa = None
        self.usuario_actual = None  # dict con usuario y rol

    def _cargar_usuarios(self):
        return [{"usuario": f[0], "password": f[1], "rol": str(f[2]).lower()} for f in self.excel.cargar_hoja(HOJA_USUARIOS)]

    def validar_credenciales(self, usr, pwd):
        for u in self.usuarios:
            if u["usuario"] == usr and u["password"] == pwd:
                return u
        return None

    def abrir_sesion(self, usuario_dict):
        self.usuario_actual = usuario_dict
        self.id_sesion_activa = self.sesion.abrir_sesion(usuario_dict["usuario"], usuario_dict["rol"])

    def cerrar_sesion(self):
        if self.id_sesion_activa is not None:
            self.sesion.cerrar_sesion(self.id_sesion_activa)
        self.id_sesion_activa = None
        self.usuario_actual = None


# INTERFAZ GRÁFICA (CustomTkinter)
class AppGUI:
    def __init__(self):
        # Estilos
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        # Sistema
        self.sys = SistemaDeInventario()

        # Ventana
        self.root = ctk.CTk()
        self.root.title("Tienda de Hilos Arcoíris")
        try:
            self.root.state("zoomed")
        except Exception:
            self.root.geometry("1200x700")

        # Frames
        self.frame_actual = None
        self._build_login()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        self.root.mainloop()

    # ---------- Layout helpers ----------
    def clear_frame(self):
        if self.frame_actual is not None:
            self.frame_actual.pack_forget()
            self.frame_actual.destroy()
        self.frame_actual = ctk.CTkFrame(self.root)
        self.frame_actual.pack(expand=True, fill="both")

    def header(self, parent, titulo="Tienda de Hilos Arcoíris"):
        encabezado = ctk.CTkFrame(parent, height=80, fg_color="#1e1e1e")
        encabezado.pack(fill="x")
        lbl = ctk.CTkLabel(encabezado, text=titulo, font=("Arial", 32, "bold"), text_color="#ffcc70")
        lbl.pack(pady=15)
        return encabezado

    def row_entry(self, parent, etiqueta, var: ctk.StringVar, width=280, **kwargs):
        cont = ctk.CTkFrame(parent, fg_color="transparent")
        cont.pack(pady=8)
        lbl = ctk.CTkLabel(cont, text=etiqueta, font=("Arial", 16))
        lbl.pack(side="left", padx=10)
        ent = ctk.CTkEntry(cont, textvariable=var, width=width, height=35, **kwargs)
        ent.pack(side="left")
        return ent

    def boton_volver(self, parent, destino):
        ctk.CTkButton(parent, text="↩ Volver", command=destino, width=250, height=40).pack(pady=10)

    # ---------- NUEVOS HELPERS: Lote/Tabla ----------
    def _init_tabla_lote(self, columns):
        """
        Crea una tabla (Treeview) para mostrar operaciones en lote con columnas dinámicas.
        columns: lista de tuplas (id_col, header, width, anchor)
        """
        self.lote = []  # cada item: {"accion": str, "datos": dict}
        frame_tabla = ctk.CTkFrame(self.frame_actual)
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

        col_ids = [c[0] for c in columns]
        self.tabla = ttk.Treeview(frame_tabla, columns=tuple(col_ids), show="headings", height=12)
        for cid, header, width, anchor in columns:
            self.tabla.heading(cid, text=header)
            self.tabla.column(cid, width=width, anchor=anchor)

        sb = ttk.Scrollbar(frame_tabla, orient="vertical", command=self.tabla.yview)
        self.tabla.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.tabla.pack(fill="both", expand=True)

    def _tabla_clear(self):
        if hasattr(self, "tabla") and self.tabla:
            self.tabla.delete(*self.tabla.get_children())

    def _tabla_add(self, values_tuple):
        if hasattr(self, "tabla") and self.tabla:
            self.tabla.insert("", "end", values=values_tuple)

    def _confirmar_lote(self):
        """
        Ejecuta el lote: intenta todo lo posible y reporta aciertos/errores.
        Muestra detalles, incluyendo el TOTAL de cada venta registrada.
        """
        if not getattr(self, "lote", None):
            messagebox.showinfo("Lote vacío", "No hay operaciones para confirmar.")
            return

        exitos = 0
        detalles = []   # log de cada operación (éxito y error)

        for op in self.lote:
            accion = op["accion"]
            d = op["datos"]
            try:
                if accion == "registrar_hilo":
                    ok, msg = self.sys.inventario.registrar_hilo_gui(
                        d["marca"], d["tipo"], d["codigo"], d["cantidad"], d["precio"], d["proveedor"]
                    )
                elif accion == "modificar_hilo":
                    ok, msg = self.sys.inventario.modificar_hilo_gui(
                        d["marca"],
                        d["codigo"],
                        nueva_marca=d.get("nueva_marca"),
                        nuevo_tipo=d.get("nuevo_tipo"),
                        nueva_cantidad=d.get("nueva_cantidad"),
                        nuevo_precio=d.get("nuevo_precio"),
                        nuevo_proveedor=d.get("nuevo_proveedor"),
                    )
                elif accion == "compra":
                    ok, msg = self.sys.inventario.registrar_compra_gui(d["marca"], d["codigo"], d["cantidad"], d["costo"])
                elif accion == "venta":
                    ok, msg = self.sys.inventario.registrar_venta_gui(d["marca"], d["codigo"], d["cantidad"])
                else:
                    ok, msg = False, f"Acción desconocida: {accion}"

                if ok:
                    exitos += 1
                    # Guardar detalles de éxito (incluye total en caso de venta)
                    detalles.append(f"{accion}: {msg}")
                else:
                    detalles.append(f"{accion}: {msg}")
            except Exception as e:
                detalles.append(f"{accion}: {e}")

        resumen = f"Operaciones exitosas: {exitos}\n\n"
        if detalles:
            resumen += "Detalles:\n- " + "\n- ".join(detalles)

        messagebox.showinfo("Resultado del lote", resumen)

        # Limpiar lote y tabla
        self.lote = []
        self._tabla_clear()

    # ---------- Pantallas ----------
    def _build_login(self):
        self.clear_frame()
        self.header(self.frame_actual, "Tienda de Hilos Arcoíris")

        login_frame = ctk.CTkFrame(self.frame_actual)
        login_frame.pack(expand=True, ipadx=20, ipady=20)

        ctk.CTkLabel(login_frame, text="Inicio de sesión", font=("Arial", 24, "bold")).pack(pady=20)

        v_user = ctk.StringVar()
        v_pass = ctk.StringVar()
        self.row_entry(login_frame, "Usuario", v_user)
        # Para ocultar contraseña, puedes usar: show="*"
        self.row_entry(login_frame, "Contraseña", v_pass)  # show="*"

        def hacer_login():
            usr = v_user.get().strip()
            pwd = v_pass.get().strip()
            u = self.sys.validar_credenciales(usr, pwd)
            if not u:
                messagebox.showerror("Error", "Usuario o contraseña incorrectos.")
                return
            self.sys.abrir_sesion(u)
            messagebox.showinfo("Bienvenido", f"Sesión iniciada como {u['rol']}.")
            self._build_menu()

        ctk.CTkButton(login_frame, text="Entrar", width=220, height=40, command=hacer_login).pack(pady=15)

    def _build_menu(self):
        self.clear_frame()
        self.header(self.frame_actual)

        tipo = "ADMINISTRADOR" if self.sys.usuario_actual["rol"] == "admin" else "EMPLEADO"
        ctk.CTkLabel(
            self.frame_actual,
            text=f"Usuario: {self.sys.usuario_actual['usuario']}  |  Rol: {tipo}",
            font=("Arial", 16, "italic"),
            text_color="lightblue",
        ).pack(pady=(12, 2))

        ctk.CTkLabel(self.frame_actual, text="MENÚ PRINCIPAL", font=("Arial", 26, "bold")).pack(pady=10)

        # Opciones según rol (ambos incluyen Búsqueda profunda (Hashing))
        if self.sys.usuario_actual["rol"] == "admin":
            opciones = [
                ("Registrar nuevo hilo", self.ui_registrar_hilo),
                ("Buscar hilo", self.ui_buscar_hilo),
                ("Búsqueda profunda (Hashing)", self.ui_busqueda_hash),
                ("Modificar información", self.ui_modificar_hilo),
                ("Eliminar hilo", self.ui_eliminar_hilo),
                ("Registrar compra / reabastecimiento", self.ui_registrar_compra),
                ("Registrar venta", self.ui_registrar_venta),
                ("Reportes y consultas", self.ui_reportes),
                ("Mostrar inventario completo", self.ui_inventario),
                ("Cerrar sesión y salir", self.on_close),
            ]
        else:
            opciones = [
                ("Registrar venta", self.ui_registrar_venta if PERMITIR_VENTA_EMPLEADO else self._no_permitido),
                ("Buscar hilo", self.ui_buscar_hilo),
                ("Búsqueda profunda (Hashing)", self.ui_busqueda_hash),
                ("Reportes y consultas", self.ui_reportes),
                ("Mostrar inventario completo", self.ui_inventario),
                ("Cerrar sesión y salir", self.on_close),
            ]

        botones = ctk.CTkFrame(self.frame_actual, fg_color="transparent")
        botones.pack(pady=10)
        for texto, cmd in opciones:
            ctk.CTkButton(botones, text=texto, command=cmd, width=420, height=42, font=("Arial", 16)).pack(pady=7)

        ctk.CTkButton(
            self.frame_actual, text="↩ Cerrar sesión", fg_color="#444", hover_color="#666", command=self._logout, width=280, height=38
        ).pack(pady=10)

    def _logout(self):
        self.sys.cerrar_sesion()
        self._build_login()

    def _no_permitido(self):
        messagebox.showwarning("Restringido", "Ventas solo permitidas al administrador.")

    # ---------- Subpantalla: Búsqueda Profunda (Hashing) ----------
    def ui_busqueda_hash(self):
        self.clear_frame()
        self.header(self.frame_actual, "Búsqueda Profunda (Hashing)")

        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=20)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()

        self.row_entry(form, "Marca", v_marca)
        self.row_entry(form, "Código de color", v_codigo)

        frame_tabla = ctk.CTkFrame(self.frame_actual)
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

        columnas = ("ID", "Marca", "Código", "Tipo", "Cantidad", "Precio", "Proveedor")
        tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=10)

        for col in columnas:
            tabla.heading(col, text=col)
            tabla.column(col, width=120, anchor="center")

        tabla.column("ID", width=70)
        tabla.column("Cantidad", width=80)
        tabla.column("Precio", width=100)

        sb = ttk.Scrollbar(frame_tabla, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        tabla.pack(fill="both", expand=True)

        def buscar_hash():
            marca = v_marca.get().strip()
            codigo = v_codigo.get().strip()

            if marca == "" or codigo == "":
                messagebox.showwarning("Atención", "Debes ingresar marca y código.")
                return

            tabla.delete(*tabla.get_children())

            res = self.sys.inventario.buscar_hash(marca, codigo)
            if res is None:
                messagebox.showinfo("Sin resultados", "No se encontró el hilo usando hashing.")
                return

            tabla.insert(
                "",
                "end",
                values=(
                    res["id"],
                    res["marca"],
                    res["codigo_color"],
                    res["descripcion"],
                    res["cantidad"],
                    f"{res['precio_unitario']:.2f}",
                    res["proveedor"],
                ),
            )

        ctk.CTkButton(form, text="Buscar (hashing)", command=buscar_hash, width=200, height=40).pack(pady=20)

        self.boton_volver(self.frame_actual, self._build_menu)

    # ---------- Subpantallas (con tablas + lote) ----------
    def ui_registrar_hilo(self):
        self.clear_frame()
        self.header(self.frame_actual, "Registrar nuevo hilo (Lote)")

        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_tipo = ctk.StringVar()
        v_codigo = ctk.StringVar()
        v_cantidad = ctk.StringVar()
        v_precio = ctk.StringVar()
        v_proveedor = ctk.StringVar()

        self.row_entry(form, "Marca", v_marca)
        self.row_entry(form, "Tipo", v_tipo)
        self.row_entry(form, "Código de color", v_codigo)
        self.row_entry(form, "Cantidad", v_cantidad)
        self.row_entry(form, "Precio por unidad", v_precio)
        self.row_entry(form, "Proveedor", v_proveedor)

        cols = [
            ("marca", "Marca", 120, "w"),
            ("codigo", "Código", 90, "center"),
            ("tipo", "Tipo", 160, "w"),
            ("cantidad", "Cantidad", 80, "center"),
            ("precio", "Precio (Q)", 100, "e"),
            ("proveedor", "Proveedor", 160, "w"),
        ]
        self._init_tabla_lote(cols)

        def agregar_lote():
            datos = {
                "marca": v_marca.get().strip(),
                "tipo": v_tipo.get().strip(),
                "codigo": v_codigo.get().strip(),
                "cantidad": v_cantidad.get().strip(),
                "precio": v_precio.get().strip(),
                "proveedor": v_proveedor.get().strip(),
            }

            if any(v == "" for v in datos.values()):
                messagebox.showwarning("Atención", "Por favor llena todos los campos.")
                return

            n_cant = Utilidades.leer_entero_str(datos["cantidad"], minimo=0)
            n_prec = Utilidades.leer_float_str(datos["precio"], minimo=0.0)
            if n_cant is None or n_prec is None:
                messagebox.showwarning("Atención", "Cantidad o precio inválido.")
                return

            # Evitar duplicado en lote por (marca, codigo)
            clave = (datos["marca"].lower(), datos["codigo"])
            for op in self.lote:
                if op["accion"] == "registrar_hilo":
                    d = op["datos"]
                    if (d["marca"].lower(), d["codigo"]) == clave:
                        messagebox.showwarning("Duplicado en lote", "Ya agregaste este Código para esa Marca en el lote.")
                        return

            self.lote.append(
                {
                    "accion": "registrar_hilo",
                    "datos": {
                        "marca": datos["marca"],
                        "tipo": datos["tipo"],
                        "codigo": datos["codigo"],
                        "cantidad": n_cant,
                        "precio": n_prec,
                        "proveedor": datos["proveedor"],
                    },
                }
            )

            self._tabla_add((datos["marca"], datos["codigo"], datos["tipo"], n_cant, f"{n_prec:.2f}", datos["proveedor"]))

            # Limpiar inputs
            v_tipo.set("")
            v_codigo.set("")
            v_cantidad.set("")
            v_precio.set("")
            v_proveedor.set("")

        btns = ctk.CTkFrame(self.frame_actual, fg_color="transparent")
        btns.pack(pady=8)
        ctk.CTkButton(btns, text="Agregar a lote", command=agregar_lote, width=200).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Confirmar lote y registrar todo", command=self._confirmar_lote, width=280).pack(
            side="left", padx=6
        )

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_buscar_hilo(self):
        self.clear_frame()
        self.header(self.frame_actual, "Buscar Hilo (Por Marca y por Código)")

        # --- FORMULARIO ---
        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()

        self.row_entry(form, "Marca", v_marca)

        # --- Tabla creada ANTES de las funciones ---
        frame_tabla = ctk.CTkFrame(self.frame_actual)
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

        columnas = ("ID", "Marca", "Código", "Tipo", "Cantidad", "Precio", "Proveedor")
        tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=12)

        for col in columnas:
            tabla.heading(col, text=col)
            tabla.column(col, width=120, anchor="center")
        tabla.column("ID", width=60)
        tabla.column("Cantidad", width=80)
        tabla.column("Precio", width=100)

        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tabla.pack(fill="both", expand=True)

        #    FUNCIONES INTERNAS
        # Campo Código
        entry_codigo = self.row_entry(form, "Código de color", v_codigo)
        entry_codigo.configure(state="disabled")

        # Botón Buscar por Código
        btn_buscar_codigo = ctk.CTkButton(form, text="Buscar por código", width=200, height=40)
        btn_buscar_codigo.pack(pady=12)
        btn_buscar_codigo.configure(state="disabled")

        def buscar_por_marca():
            marca = v_marca.get().strip()
            if marca == "":
                messagebox.showwarning("Atención", "Ingrese una marca.")
                return

            tabla.delete(*tabla.get_children())

            encontrados = [
                h for h in self.sys.inventario.reporte_inventario() if h["marca"].lower() == marca.lower()
            ]

            if not encontrados:
                messagebox.showinfo("Sin resultados", "No se encontraron hilos para esta marca.")
                entry_codigo.configure(state="disabled")
                btn_buscar_codigo.configure(state="disabled")
                return

            for h in encontrados:
                tabla.insert(
                    "", "end",
                    values=(
                        h["id"],
                        h["marca"],
                        h["codigo_color"],
                        h["descripcion"],
                        h["cantidad"],
                        f"{h['precio_unitario']:.2f}",
                        h["proveedor"],
                    ),
                )

            entry_codigo.configure(state="normal")
            btn_buscar_codigo.configure(state="normal")

        ctk.CTkButton(
            form, text="Mostrar inventario de la marca", width=280, height=40, command=buscar_por_marca
        ).pack(pady=10)

        def buscar_por_codigo():
            marca = v_marca.get().strip()
            codigo = v_codigo.get().strip()

            if codigo == "":
                messagebox.showwarning("Atención", "Ingrese el código de color.")
                return

            tabla.delete(*tabla.get_children())

            h = self.sys.inventario.obtener_por_marca_codigo(marca, codigo)
            if h:
                tabla.insert(
                    "", "end",
                    values=(
                        h["id"],
                        h["marca"],
                        h["codigo_color"],
                        h["descripcion"],
                        h["cantidad"],
                        f"{h['precio_unitario']:.2f}",
                        h["proveedor"],
                    ),
                )
            else:
                messagebox.showinfo("Sin resultados", "No se encontró ese código para esta marca.")

        btn_buscar_codigo.configure(command=buscar_por_codigo)

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_modificar_hilo(self):
        self.clear_frame()
        self.header(self.frame_actual, "Modificar Hilo (Lote)")

        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()
        v_nueva_marca = ctk.StringVar()
        v_nuevo_tipo = ctk.StringVar()
        v_nueva_cantidad = ctk.StringVar()
        v_nuevo_precio = ctk.StringVar()
        v_nuevo_prov = ctk.StringVar()

        self.row_entry(form, "Marca (actual)", v_marca)
        self.row_entry(form, "Código de color (actual)", v_codigo)
        self.row_entry(form, "Nueva marca (opcional)", v_nueva_marca)
        self.row_entry(form, "Nuevo tipo (opcional)", v_nuevo_tipo)
        self.row_entry(form, "Nueva cantidad (opcional)", v_nueva_cantidad)
        self.row_entry(form, "Nuevo precio (opcional)", v_nuevo_precio)
        self.row_entry(form, "Nuevo proveedor (opcional)", v_nuevo_prov)

        cols = [
            ("marca", "Marca (act.)", 120, "w"),
            ("codigo", "Código (act.)", 100, "center"),
            ("nmarca", "Nueva Marca", 120, "w"),
            ("ntipo", "Nuevo Tipo", 140, "w"),
            ("ncant", "Nueva Cant.", 100, "center"),
            ("nprec", "Nuevo Precio", 110, "e"),
            ("nprov", "Nuevo Proveedor", 150, "w"),
        ]
        self._init_tabla_lote(cols)

        def agregar_lote():
            m = v_marca.get().strip()
            c = v_codigo.get().strip()
            if m == "" or c == "":
                messagebox.showwarning("Atención", "Marca y Código actuales son obligatorios.")
                return

            nc = v_nueva_cantidad.get().strip()
            np = v_nuevo_precio.get().strip()
            val_nc = None if nc == "" else Utilidades.leer_entero_str(nc, minimo=0)
            val_np = None if np == "" else Utilidades.leer_float_str(np, minimo=0.0)
            if nc != "" and val_nc is None:
                messagebox.showwarning("Atención", "Nueva cantidad inválida.")
                return
            if np != "" and val_np is None:
                messagebox.showwarning("Atención", "Nuevo precio inválido.")
                return

            datos = {
                "marca": m,
                "codigo": c,
                "nueva_marca": v_nueva_marca.get().strip() or None,
                "nuevo_tipo": v_nuevo_tipo.get().strip() or None,
                "nueva_cantidad": None if nc == "" else nc,
                "nuevo_precio": None if np == "" else np,
                "nuevo_proveedor": v_nuevo_prov.get().strip() or None,
            }

            self.lote.append({"accion": "modificar_hilo", "datos": datos})
            self._tabla_add(
                (
                    datos["marca"],
                    datos["codigo"],
                    datos["nueva_marca"] or "",
                    datos["nuevo_tipo"] or "",
                    datos["nueva_cantidad"] or "",
                    datos["nuevo_precio"] or "",
                    datos["nuevo_proveedor"] or "",
                )
            )

            # Limpiar inputs opcionales
            v_nueva_marca.set("")
            v_nuevo_tipo.set("")
            v_nueva_cantidad.set("")
            v_nuevo_precio.set("")
            v_nuevo_prov.set("")

        btns = ctk.CTkFrame(self.frame_actual, fg_color="transparent")
        btns.pack(pady=8)
        ctk.CTkButton(btns, text="Agregar modificación al lote", command=agregar_lote, width=260).pack(
            side="left", padx=6
        )
        ctk.CTkButton(btns, text="Confirmar lote y aplicar cambios", command=self._confirmar_lote, width=280).pack(
            side="left", padx=6
        )

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_eliminar_hilo(self):
        self.clear_frame()
        self.header(self.frame_actual, "Eliminar Hilo")
        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()
        self.row_entry(form, "Marca", v_marca)
        self.row_entry(form, "Código de color", v_codigo)

        def confirmar():
            ok, msg = self.sys.inventario.eliminar_hilo_gui(v_marca.get().strip(), v_codigo.get().strip())
            if ok:
                messagebox.showinfo("Éxito", msg)
                self._build_menu()
            else:
                messagebox.showwarning("Atención", msg)

        ctk.CTkButton(form, text="Eliminar hilo", command=confirmar, width=240, height=40).pack(pady=12)
        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_registrar_compra(self):
        self.clear_frame()
        self.header(self.frame_actual, "Registrar Compra / Reabastecimiento (Lote)")

        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()
        v_cantidad = ctk.StringVar()
        v_costo = ctk.StringVar()

        self.row_entry(form, "Marca", v_marca)
        self.row_entry(form, "Código de color", v_codigo)
        self.row_entry(form, "Cantidad comprada", v_cantidad)
        self.row_entry(form, "Costo por unidad", v_costo)

        cols = [("marca", "Marca", 120, "w"), ("codigo", "Código", 90, "center"), ("cantidad", "Cantidad", 90, "center"), ("costo", "Costo (Q)", 100, "e")]
        self._init_tabla_lote(cols)

        def agregar_lote():
            m = v_marca.get().strip()
            c = v_codigo.get().strip()
            cant = Utilidades.leer_entero_str(v_cantidad.get().strip(), minimo=1)
            costo = Utilidades.leer_float_str(v_costo.get().strip(), minimo=0.0)

            if any(x in (None, "") for x in [m, c, cant, costo]):
                messagebox.showwarning("Atención", "Complete los datos correctamente.")
                return

            self.lote.append({"accion": "compra", "datos": {"marca": m, "codigo": c, "cantidad": cant, "costo": costo}})
            self._tabla_add((m, c, cant, f"{costo:.2f}"))

            v_codigo.set("")
            v_cantidad.set("")
            v_costo.set("")

        btns = ctk.CTkFrame(self.frame_actual, fg_color="transparent")
        btns.pack(pady=8)
        ctk.CTkButton(btns, text="Agregar compra al lote", command=agregar_lote, width=240).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Confirmar lote y registrar compras", command=self._confirmar_lote, width=280).pack(
            side="left", padx=6
        )

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_registrar_venta(self):
        self.clear_frame()
        self.header(self.frame_actual, "Registrar Venta (Lote)")

        form = ctk.CTkFrame(self.frame_actual)
        form.pack(pady=10)

        v_marca = ctk.StringVar()
        v_codigo = ctk.StringVar()
        v_cantidad = ctk.StringVar()

        self.row_entry(form, "Marca", v_marca)
        self.row_entry(form, "Código de color", v_codigo)
        self.row_entry(form, "Cantidad vendida", v_cantidad)

        cols = [("marca", "Marca", 120, "w"), ("codigo", "Código", 90, "center"), ("cantidad", "Cantidad", 90, "center")]
        self._init_tabla_lote(cols)

        def agregar_lote():
            m = v_marca.get().strip()
            c = v_codigo.get().strip()
            cant = Utilidades.leer_entero_str(v_cantidad.get().strip(), minimo=1)
            if any(x in (None, "") for x in [m, c, cant]):
                messagebox.showwarning("Atención", "Complete los datos correctamente.")
                return

            self.lote.append({"accion": "venta", "datos": {"marca": m, "codigo": c, "cantidad": cant}})
            self._tabla_add((m, c, cant))

            v_codigo.set("")
            v_cantidad.set("")

        btns = ctk.CTkFrame(self.frame_actual, fg_color="transparent")
        btns.pack(pady=8)
        ctk.CTkButton(btns, text="Agregar venta al lote", command=agregar_lote, width=220).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Confirmar lote y registrar ventas", command=self._confirmar_lote, width=280).pack(
            side="left", padx=6
        )

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_reportes(self):
        self.clear_frame()
        self.header(self.frame_actual, "Reportes y Consultas")

        cont = ctk.CTkFrame(self.frame_actual)
        cont.pack(pady=10)

        frame_tabla = ctk.CTkFrame(self.frame_actual)
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

        columnas = ("Col1", "Col2", "Col3", "Col4", "Col5", "Col6")
        tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=15)
        for col in columnas:
            tabla.heading(col, text=col)
            tabla.column(col, width=150, anchor="center")
        scrollbar = ttk.Scrollbar(frame_tabla, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tabla.pack(fill="both", expand=True)

        def rep_inv():
            tabla.delete(*tabla.get_children())
            tabla.heading("Col1", text="ID")
            tabla.heading("Col2", text="Marca")
            tabla.heading("Col3", text="Código")
            tabla.heading("Col4", text="Tipo")
            tabla.heading("Col5", text="Cantidad")
            tabla.heading("Col6", text="Precio (Q)")

            data = self.sys.inventario.reporte_inventario()
            for h in data:
                tabla.insert(
                    "",
                    "end",
                    values=(h["id"], h["marca"], h["codigo_color"], h["descripcion"], h["cantidad"], f"{h['precio_unitario']:.2f}"),
                )

        def rep_ven():
            tabla.delete(*tabla.get_children())
            tabla.heading("Col1", text="Marca")
            tabla.heading("Col2", text="Código")
            tabla.heading("Col3", text="Tipo")
            tabla.heading("Col4", text="Cantidad")
            tabla.heading("Col5", text="Total (Q)")
            tabla.heading("Col6", text="")

            data = self.sys.inventario.reporte_ventas()
            for v in data:
                tabla.insert(
                    "",
                    "end",
                    values=(v["marca"], v["codigo_color"], v["descripcion"], v["cantidad"], f"{v['total']:.2f}", ""),
                )

        def rep_com():
            tabla.delete(*tabla.get_children())
            tabla.heading("Col1", text="Marca")
            tabla.heading("Col2", text="Código")
            tabla.heading("Col3", text="Tipo")
            tabla.heading("Col4", text="Cantidad")
            tabla.heading("Col5", text="Costo (Q)")
            tabla.heading("Col6", text="Total (Q)")

            data = self.sys.inventario.reporte_compras()
            for c in data:
                tabla.insert(
                    "",
                    "end",
                    values=(c["marca"], c["codigo_color"], c["descripcion"], c["cantidad"], f"{c['costo_unitario']:.2f}", f"{c['total']:.2f}"),
                )

        btns = ctk.CTkFrame(cont, fg_color="transparent")
        btns.pack(pady=6)
        ctk.CTkButton(btns, text="Inventario", command=rep_inv, width=180, height=34).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Historial Ventas", command=rep_ven, width=180, height=34).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Historial Compras", command=rep_com, width=180, height=34).pack(side="left", padx=6)

        self.boton_volver(self.frame_actual, self._build_menu)

    def ui_inventario(self):
        self.clear_frame()
        self.header(self.frame_actual, "Inventario Completo")

        frame_tabla = ctk.CTkFrame(self.frame_actual)
        frame_tabla.pack(fill="both", expand=True, padx=20, pady=20)

        columnas = ("ID", "Marca", "Código", "Tipo", "Cantidad", "Precio", "Proveedor")
        tabla = ttk.Treeview(frame_tabla, columns=columnas, show="headings", height=20)

        for col in columnas:
            tabla.heading(col, text=col)
            tabla.column(col, width=120, anchor="center")

        tabla.column("ID", width=50)
        tabla.column("Cantidad", width=80)
        tabla.column("Precio", width=100)

        scrollbar_y = ttk.Scrollbar(frame_tabla, orient="vertical", command=tabla.yview)
        tabla.configure(yscrollcommand=scrollbar_y.set)
        scrollbar_y.pack(side="right", fill="y")
        tabla.pack(fill="both", expand=True)

        data = self.sys.inventario.reporte_inventario()
        for h in data:
            tabla.insert(
                "",
                "end",
                values=(h["id"], h["marca"], h["codigo_color"], h["descripcion"], h["cantidad"], f"{h['precio_unitario']:.2f}", h["proveedor"]),
            )

        self.boton_volver(self.frame_actual, self._build_menu)

    # ---------- cierre ----------
    def on_close(self):
        # Cierra sesión si estaba abierta
        try:
            self.sys.cerrar_sesion()
        except Exception:
            pass
        self.root.destroy()


# MAIN
if __name__ == "__main__":
    AppGUI()
